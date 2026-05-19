import streamlit as st
from pathlib import Path
from supabase import create_client
import base64
from datetime import datetime

# ==============================
# RUTAS DE ARCHIVOS
# ==============================
BASE_DIR = Path(__file__).resolve().parent
LOGO = BASE_DIR / "logo.png"
FONDO_LOGIN = BASE_DIR / "fondo_login.jpg"

# ==============================
# CONFIGURACIÓN DE PÁGINA
# ==============================
st.set_page_config(
    page_title="Registro de cementerios",
    page_icon="🪦",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ==============================
# SUPABASE
# ==============================
@st.cache_resource
def get_supabase():
    return create_client(
        st.secrets["SUPABASE_URL"],
        st.secrets["SUPABASE_KEY"]
    )

supabase = get_supabase()

# ==============================
# ESTADO DE SESIÓN
# ==============================
if "user" not in st.session_state:
    st.session_state.user = None

if "user_role" not in st.session_state:
    st.session_state.user_role = None

if "user_nombre" not in st.session_state:
    st.session_state.user_nombre = None

if "username" not in st.session_state:
    st.session_state.username = None

if "editando_id" not in st.session_state:
    st.session_state.editando_id = None

# ==============================
# DATOS FIJOS
# ==============================
ESTADOS = ["Provisional", "Aceptado", "Certificado"]

TERRITORIOS = {
    "Ahuachapán": {
        "Ahuachapán Norte": ["Atiquizaya", "El Refugio", "San Lorenzo", "Turín"],
        "Ahuachapán Centro": ["Ahuachapán", "Apaneca", "Concepción de Ataco", "Tacuba"],
        "Ahuachapán Sur": ["Guaymango", "Jujutla", "San Francisco Menéndez", "San Pedro Puxtla"],
    },
    "Santa Ana": {
        "Santa Ana Norte": ["Masahuat", "Metapán", "Santa Rosa Guachipilín", "Texistepeque"],
        "Santa Ana Centro": ["Santa Ana"],
        "Santa Ana Este": ["Coatepeque", "El Congo"],
        "Santa Ana Oeste": ["Candelaria de la Frontera", "Chalchuapa", "El Porvenir", "San Antonio Pajonal", "San Sebastián Salitrillo", "Santiago de la Frontera"],
    },
    "Sonsonate": {
        "Sonsonate Norte": ["Juayúa", "Nahuizalco", "Salcoatitán", "Santa Catarina Masahuat"],
        "Sonsonate Centro": ["Izalco", "Nahulingo", "San Antonio del Monte", "Sonsonate", "Sonzacate"],
        "Sonsonate Este": ["Armenia", "Caluco", "Cuisnahuat", "Santa Isabel Ishuatán", "Santo Domingo de Guzmán"],
        "Sonsonate Oeste": ["Acajutla", "San Julián"],
    },
    "Chalatenango": {
        "Chalatenango Norte": ["Citalá", "La Palma", "San Ignacio"],
        "Chalatenango Centro": ["Agua Caliente", "Dulce Nombre de María", "El Paraíso", "La Reina", "Nueva Concepción", "San Fernando", "San Francisco Morazán", "San Rafael", "Santa Rita", "Tejutla"],
        "Chalatenango Sur": ["Arcatao", "Azacualpa", "Chalatenango", "Comalapa", "Concepción Quezaltepeque", "El Carrizal", "La Laguna", "Las Vueltas", "Nombre de Jesús", "Nueva Trinidad", "Ojos de Agua", "Potonico", "San Antonio de la Cruz", "San Antonio Los Ranchos", "San Francisco Lempa", "San Isidro Labrador", "San José Cancasque", "San José Las Flores", "San Luis del Carmen", "San Miguel de Mercedes"],
    },
    "La Libertad": {
        "La Libertad Norte": ["Quezaltepeque", "San Matías", "San Pablo Tacachico"],
        "La Libertad Centro": ["Ciudad Arce", "San Juan Opico"],
        "La Libertad Oeste": ["Colón", "Jayaque", "Sacacoyo", "Talnique", "Tepecoyo"],
        "La Libertad Este": ["Antiguo Cuscatlán", "Huizúcar", "Nuevo Cuscatlán", "San José Villanueva", "Zaragoza"],
        "La Libertad Costa": ["Chiltiupán", "Jicalapa", "La Libertad", "Tamanique", "Teotepeque"],
        "La Libertad Sur": ["Comasagua", "Santa Tecla"],
    },
    "San Salvador": {
        "San Salvador Norte": ["Aguilares", "El Paisnal", "Guazapa"],
        "San Salvador Oeste": ["Apopa", "Nejapa"],
        "San Salvador Este": ["Ilopango", "San Martín", "Soyapango", "Tonacatepeque"],
        "San Salvador Centro": ["Ayutuxtepeque", "Ciudad Delgado", "Cuscatancingo", "Mejicanos", "San Salvador"],
        "San Salvador Sur": ["Panchimalco", "Rosario de Mora", "San Marcos", "Santiago Texacuangos", "Santo Tomás"],
    },
    "Cuscatlán": {
        "Cuscatlán Norte": ["Oratorio de Concepción", "San Bartolomé Perulapía", "San José Guayabal", "San Pedro Perulapán", "Suchitoto"],
        "Cuscatlán Sur": ["Candelaria", "Cojutepeque", "El Carmen", "El Rosario", "Monte San Juan", "San Cristóbal", "San Rafael Cedros", "San Ramón", "Santa Cruz Analquito", "Santa Cruz Michapa", "Tenancingo"],
    },
    "La Paz": {
        "La Paz Oeste": ["Cuyultitán", "Olocuilta", "San Francisco Chinameca", "San Juan Talpa", "San Luis Talpa", "San Pedro Masahuat", "Tapalhuaca"],
        "La Paz Centro": ["El Rosario", "Jerusalén", "Mercedes La Ceiba", "Paraíso de Osorio", "San Antonio Masahuat", "San Emigdio", "San Juan Tepezontes", "San Luis La Herradura", "San Miguel Tepezontes", "San Pedro Nonualco", "Santa María Ostuma", "Santiago Nonualco"],
        "La Paz Este": ["San Juan Nonualco", "San Rafael Obrajuelo", "Zacatecoluca"],
    },
    "Cabañas": {
        "Cabañas Este": ["Dolores", "Guacotecti", "San Isidro", "Sensuntepeque", "Victoria"],
        "Cabañas Oeste": ["Cinquera", "Ilobasco", "Jutiapa", "Tejutepeque"],
    },
    "San Vicente": {
        "San Vicente Norte": ["Apastepeque", "San Cayetano Istepeque", "San Esteban Catarina", "San Ildefonso", "San Lorenzo", "San Sebastián", "Santa Clara"],
        "San Vicente Sur": ["Guadalupe", "San Vicente", "Santo Domingo", "Tecoluca", "Tepetitán", "Verapaz"],
    },
    "Usulután": {
        "Usulután Norte": ["Alegría", "Berlín", "El Triunfo", "Estanzuelas", "Jucuapa", "Mercedes Umaña", "Nueva Granada", "San Buenaventura", "Santiago de María", "Tecapán"],
        "Usulután Este": ["California", "Concepción Batres", "Ereguayquín", "Jiquilisco", "Jucuarán", "Ozatlán", "Puerto El Triunfo", "San Agustín", "San Dionisio", "San Francisco Javier", "Santa Elena", "Santa María", "Tecapán", "Usulután"],
        "Usulután Oeste": ["Jiquilisco", "Ozatlán", "San Dionisio", "Santa María"],
    },
    "San Miguel": {
        "San Miguel Norte": ["Carolina", "Chapeltique", "Ciudad Barrios", "Nuevo Edén de San Juan", "San Antonio del Mosco", "San Gerardo", "San Luis de la Reina", "Sesori"],
        "San Miguel Centro": ["Chinameca", "Chirilagua", "Comacarán", "Moncagua", "Quelepa", "San Miguel", "Uluazapa"],
        "San Miguel Oeste": ["El Tránsito", "Lolotique", "Nueva Guadalupe", "San Jorge", "San Rafael Oriente"],
    },
    "Morazán": {
        "Morazán Norte": ["Arambala", "Cacaopera", "Corinto", "El Rosario", "Joateca", "Jocoaitique", "Meanguera", "Perquín", "San Fernando", "San Isidro", "Torola"],
        "Morazán Sur": ["Chilanga", "Delicias de Concepción", "El Divisadero", "Gualococti", "Guatajiagua", "Jocoro", "Lolotiquillo", "Osicala", "San Carlos", "San Francisco Gotera", "San Simón", "Sensembra", "Sociedad", "Yamabal", "Yoloaiquín"],
    },
    "La Unión": {
        "La Unión Norte": ["Anamorós", "Bolívar", "Concepción de Oriente", "El Sauce", "Lislique", "Nueva Esparta", "Pasaquina", "Polorós", "San José La Fuente", "Santa Rosa de Lima"],
        "La Unión Sur": ["Conchagua", "El Carmen", "Intipucá", "La Unión", "Meanguera del Golfo", "San Alejo", "Yayantique", "Yucuaiquín"],
    }
}

# ==============================
# UTILIDADES
# ==============================
def imagen_a_base64(ruta):
    if ruta.exists():
        return base64.b64encode(ruta.read_bytes()).decode()
    return ""

def es_url_valida(url):
    if not str(url).strip():
        return True
    url = str(url).strip().lower()
    return url.startswith("http://") or url.startswith("https://")

def es_coordenada_valida(texto):
    if not str(texto).strip():
        return True
    partes = str(texto).split(",")
    if len(partes) != 2:
        return False
    try:
        lat = float(partes[0].strip())
        lon = float(partes[1].strip())
        return -90 <= lat <= 90 and -180 <= lon <= 180
    except ValueError:
        return False

def validar_usuario(usuario, contrasena):
    try:
        response = (
            supabase.table("usuarios_acceso")
            .select("*")
            .eq("usuario", usuario.strip())
            .eq("contrasena", contrasena.strip())
            .eq("activo", True)
            .execute()
        )

        if response.data and len(response.data) > 0:
            fila = response.data[0]
            return {
                "id": fila["id"],
                "usuario": fila["usuario"],
                "rol": fila["rol"],
                "nombre": fila.get("nombre_mostrar", fila["usuario"])
            }
        return None
    except Exception as e:
        st.error(f"Error al validar usuario: {e}")
        return None

def guardar_cementerio(datos):
    try:
        response = supabase.table("cementerios").insert(datos).execute()
        return True, response
    except Exception as e:
        return False, str(e)

def actualizar_cementerio(id_registro, datos):
    try:
        response = supabase.table("cementerios").update(datos).eq("id", id_registro).execute()
        return True, response
    except Exception as e:
        return False, str(e)

def eliminar_cementerio(id_registro):
    try:
        supabase.table("cementerios").delete().eq("id", id_registro).execute()
        return True
    except Exception:
        return False

def cargar_cementerios():
    try:
        response = supabase.table("cementerios").select("*").order("created_at", desc=True).execute()
        return response.data if response.data else []
    except Exception:
        return []

def cargar_cementerio_por_id(id_registro):
    try:
        response = supabase.table("cementerios").select("*").eq("id", id_registro).execute()
        if response.data and len(response.data) > 0:
            return response.data[0]
        return None
    except Exception:
        return None

# ==============================
# DESCARGA A EXCEL
# ==============================
def generar_excel(datos):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Cementerios"

    encabezados = [
        "ID", "Nombre", "Departamento", "Municipio", "Distrito",
        "Estado", "Coordenadas", "Google Maps", "Enlace 1",
        "Enlace 2", "Enlace 3", "Foto URL", "Observaciones",
        "Aprobado QA", "Comentarios QA", "Revisado por", "Fecha Revisión"
    ]

    for col, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=encabezado)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="5A3F8C", end_color="5A3F8C", fill_type="solid")
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for fila_idx, registro in enumerate(datos, 2):
        ws.cell(row=fila_idx, column=1, value=registro.get("codigo_cementerio"))
        ws.cell(row=fila_idx, column=2, value=registro.get("nombre_oficial"))
        ws.cell(row=fila_idx, column=3, value=registro.get("departamento"))
        ws.cell(row=fila_idx, column=4, value=registro.get("municipio"))
        ws.cell(row=fila_idx, column=5, value=registro.get("distrito"))
        ws.cell(row=fila_idx, column=6, value=registro.get("estado"))
        ws.cell(row=fila_idx, column=7, value=registro.get("coordenadas"))
        ws.cell(row=fila_idx, column=8, value=registro.get("google_maps_url"))
        ws.cell(row=fila_idx, column=9, value=registro.get("enlace_principal"))
        ws.cell(row=fila_idx, column=10, value=registro.get("enlace_2"))
        ws.cell(row=fila_idx, column=11, value=registro.get("enlace_3"))
        ws.cell(row=fila_idx, column=12, value=registro.get("foto_url"))
        ws.cell(row=fila_idx, column=13, value=registro.get("observaciones"))
        ws.cell(row=fila_idx, column=14, value="Sí" if registro.get("aprobado_qa") else "No")
        ws.cell(row=fila_idx, column=15, value=registro.get("comentarios_qa"))
        ws.cell(row=fila_idx, column=16, value=registro.get("revisado_por_qa"))
        ws.cell(row=fila_idx, column=17, value=registro.get("fecha_revision_qa"))

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# ==============================
# ESTILOS LOGIN
# ==============================
def aplicar_estilos_login():
    fondo_b64 = imagen_a_base64(FONDO_LOGIN)

    st.markdown(
        f"""
        <style>
        #MainMenu {{
            visibility: hidden;
        }}

        header {{
            visibility: hidden;
        }}

        footer {{
            visibility: hidden;
        }}

        [data-testid="stToolbar"],
        [data-testid="stDecoration"],
        [data-testid="stStatusWidget"] {{
            display: none !important;
        }}

        .stApp {{
            background:
                linear-gradient(rgba(0,0,0,0.18), rgba(0,0,0,0.18)),
                url("data:image/jpg;base64,{fondo_b64}");
            background-size: cover;
            background-position: center;
            background-repeat: no-repeat;
            background-attachment: fixed;
        }}

        [data-testid="stAppViewContainer"] > .main {{
            min-height: 100vh;
            display: flex;
            justify-content: center;
            align-items: center;
            background: rgba(255,255,255,0.02);
        }}

        .block-container {{
            max-width: 1200px !important;
            width: 100% !important;
            padding-top: 0 !important;
            padding-bottom: 0 !important;
            padding-left: 1rem !important;
            padding-right: 1rem !important;
        }}

        div[data-testid="stForm"] {{
            background: rgba(255, 255, 255, 0.95);
            border: 1px solid rgba(255,255,255,0.72);
            border-radius: 24px;
            padding: 36px 42px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.16);
            backdrop-filter: blur(4px);
            max-width: 520px;
            margin: 0 auto;
        }}

        [data-testid="stImage"] {{
            text-align: center;
            margin-bottom: 20px;
        }}

        [data-testid="stImage"] img {{
            display: block;
            margin-left: auto;
            margin-right: auto;
            width: 260px !important;
            max-width: 260px !important;
            height: auto !important;
        }}

        .login-title {{
            text-align: center;
            color: #2f2544;
            font-size: 2.15rem;
            font-weight: 700;
            margin-top: 14px;
            margin-bottom: 10px;
            line-height: 1.2;
        }}

        .login-subtitle {{
            text-align: center;
            color: #5f6673;
            font-size: 1rem;
            margin-bottom: 6px;
            line-height: 1.5;
        }}

        .login-author {{
            text-align: center;
            color: #7b8089;
            font-size: 0.95rem;
            margin-bottom: 1.15rem;
        }}

        [data-testid="stTextInput"] label p {{
            color: #384252 !important;
            font-weight: 600 !important;
            font-size: 0.96rem !important;
        }}

        [data-testid="stTextInput"] input {{
            background: #ffffff !important;
            color: #243041 !important;
            border: 1px solid #d8dce3 !important;
            border-radius: 14px !important;
            padding-top: 0.90rem !important;
            padding-bottom: 0.90rem !important;
            padding-left: 1rem !important;
            padding-right: 45px !important;
            font-size: 1rem !important;
        }}

        [data-testid="stTextInput"] input:focus {{
            border: 1px solid #6b4fa1 !important;
            box-shadow: 0 0 0 2px rgba(107,79,161,0.14) !important;
        }}

        [data-testid="stFormSubmitButton"] button {{
            width: 100% !important;
            border: none !important;
            border-radius: 14px !important;
            background: #5a3f8c !important;
            color: white !important;
            font-weight: 700 !important;
            font-size: 1rem !important;
            padding: 0.95rem 1rem !important;
            margin-top: 0.7rem !important;
        }}

        [data-testid="stFormSubmitButton"] button:hover {{
            background: #4a3276 !important;
        }}

        @media (max-width: 768px) {{
            .block-container {{
                padding-left: 0.85rem !important;
                padding-right: 0.85rem !important;
            }}

            div[data-testid="stForm"] {{
                padding: 30px 22px;
                border-radius: 20px;
            }}

            .login-title {{
                font-size: 1.75rem;
                margin-top: 16px;
            }}

            [data-testid="stImage"] img {{
                width: 210px !important;
                max-width: 210px !important;
            }}
        }}
        </style>
        """,
        unsafe_allow_html=True
    )

# ==============================
# ESTILOS APP
# ==============================
def aplicar_estilos_app():
    st.markdown(
        """
        <style>
        .stApp {
            background: #f5f3ef !important;
        }

        [data-testid="stAppViewContainer"] > .main {
            display: block !important;
            min-height: auto !important;
            background: transparent !important;
        }

        .block-container {
            max-width: 1280px !important;
            padding-top: 1.2rem !important;
            padding-bottom: 2rem !important;
        }

        .rol-badge {
            display: inline-block;
            padding: 0.25rem 0.75rem;
            border-radius: 999px;
            font-size: 0.85rem;
            font-weight: 700;
            margin-left: 0.5rem;
        }

        .badge-admin {
            background: #5a3f8c;
            color: white;
        }

        .badge-colaborador {
            background: #2d7a4f;
            color: white;
        }

        .badge-consultor {
            background: #1e5a7d;
            color: white;
        }

        .badge-qa {
            background: #d97706;
            color: white;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

# ==============================
# LOGIN
# ==============================
def mostrar_login():
    aplicar_estilos_login()

    izquierda, centro, derecha = st.columns([1.1, 1.2, 1.1])

    with centro:
        with st.form("login_form", clear_on_submit=False):
            if LOGO.exists():
                st.image(str(LOGO), width=260)

            st.markdown(
                '<div class="login-title">Registro de cementerios</div>',
                unsafe_allow_html=True
            )
            st.markdown(
                '<div class="login-subtitle">Sistema de captura, consulta y actualización de información</div>',
                unsafe_allow_html=True
            )
            st.markdown(
                '<div class="login-author">Proyecto de Claudia Barahona</div>',
                unsafe_allow_html=True
            )

            usuario = st.text_input(
                "Usuario",
                placeholder="admin, colaborador1, colaborador2, colaborador3, qa, consultor"
            )

            password = st.text_input(
                "Contraseña",
                type="password",
                placeholder="Ingrese su contraseña"
            )

            submitted = st.form_submit_button("Iniciar sesión", use_container_width=True)

            if submitted:
                if not usuario or not password:
                    st.warning("Complete usuario y contraseña.")
                else:
                    user_data = validar_usuario(usuario, password)

                    if user_data:
                        st.session_state.user = user_data["id"]
                        st.session_state.username = user_data["usuario"]
                        st.session_state.user_role = user_data["rol"]
                        st.session_state.user_nombre = user_data["nombre"]
                        st.rerun()
                    else:
                        st.error("Usuario o contraseña incorrectos.")

# ==============================
# FORMULARIO NUEVO REGISTRO
# ==============================
def mostrar_formulario_cementerio():
    st.subheader("Formulario de registro")

    departamentos = list(TERRITORIOS.keys())
    departamento = st.selectbox("Departamento", departamentos, key="departamento_select")

    municipios = list(TERRITORIOS[departamento].keys())
    municipio = st.selectbox("Municipio", municipios, key="municipio_select")

    distritos = TERRITORIOS[departamento][municipio]
    distrito = st.selectbox("Distrito", distritos, key="distrito_select")

    with st.form("form_cementerio", clear_on_submit=False):
        col1, col2 = st.columns(2)

        with col1:
            codigo_cementerio = st.text_input("ID del cementerio", placeholder="Ej. CEM-001")
            pais = st.text_input("País", value="El Salvador", disabled=True)
            nombre_oficial = st.text_input("Nombre del cementerio")
            coordenadas = st.text_input("Coordenadas", placeholder="13.7010,-89.2240")
            estado = st.selectbox("Estado", ESTADOS)
            google_maps_url = st.text_input("Enlace Google Maps", placeholder="https://...")

        with col2:
            enlace_principal = st.text_input("Enlace 1", placeholder="https://...")
            enlace_2 = st.text_input("Enlace 2", placeholder="https://...")
            enlace_3 = st.text_input("Enlace 3", placeholder="https://...")
            foto_url = st.text_input("Enlace Fotografía", placeholder="https://...")
            observaciones = st.text_area("Observaciones", height=140)

        guardar = st.form_submit_button("Guardar registro", use_container_width=True)

        if guardar:
            if not codigo_cementerio.strip():
                st.warning("Debe ingresar el ID del cementerio.")
                return

            if not nombre_oficial.strip():
                st.warning("Debe ingresar el nombre del cementerio.")
                return

            if not es_coordenada_valida(coordenadas):
                st.warning("Las coordenadas deben tener formato latitud,longitud.")
                return

            if not es_url_valida(google_maps_url):
                st.warning("El enlace de Google Maps debe empezar con http:// o https://")
                return

            if not es_url_valida(enlace_principal):
                st.warning("El Enlace 1 debe empezar con http:// o https://")
                return

            if not es_url_valida(enlace_2):
                st.warning("El Enlace 2 debe empezar con http:// o https://")
                return

            if not es_url_valida(enlace_3):
                st.warning("El Enlace 3 debe empezar con http:// o https://")
                return

            if not es_url_valida(foto_url):
                st.warning("El Enlace Fotografía debe empezar con http:// o https://")
                return

            datos = {
                "codigo_cementerio": codigo_cementerio.strip(),
                "nombre_oficial": nombre_oficial.strip(),
                "foto_url": foto_url.strip(),
                "coordenadas": coordenadas.strip(),
                "google_maps_url": google_maps_url.strip(),
                "estado": estado,
                "pais": "El Salvador",
                "departamento": departamento,
                "municipio": municipio,
                "distrito": distrito,
                "enlace_principal": enlace_principal.strip(),
                "enlace_2": enlace_2.strip(),
                "enlace_3": enlace_3.strip(),
                "observaciones": observaciones.strip(),
                "created_by": st.session_state.user,
                "created_at": datetime.now().isoformat(),
                "aprobado_qa": False
            }

            ok, resultado = guardar_cementerio(datos)

            if ok:
                st.success("Registro guardado correctamente.")
            else:
                st.error(f"No se pudo guardar el registro: {resultado}")

# ==============================
# FORMULARIO EDITAR REGISTRO
# ==============================
def mostrar_formulario_editar(registro):
    st.subheader("Editar registro")

    departamentos = list(TERRITORIOS.keys())
    idx_dep = departamentos.index(registro["departamento"]) if registro["departamento"] in departamentos else 0
    departamento = st.selectbox("Departamento", departamentos, index=idx_dep, key="edit_departamento_select")

    municipios = list(TERRITORIOS[departamento].keys())
    idx_mun = municipios.index(registro["municipio"]) if registro["municipio"] in municipios else 0
    municipio = st.selectbox("Municipio", municipios, index=idx_mun, key="edit_municipio_select")

    distritos = TERRITORIOS[departamento][municipio]
    idx_dist = distritos.index(registro["distrito"]) if registro["distrito"] in distritos else 0
    distrito = st.selectbox("Distrito", distritos, index=idx_dist, key="edit_distrito_select")

    with st.form("form_editar_cementerio", clear_on_submit=False):
        col1, col2 = st.columns(2)

        with col1:
            codigo_cementerio = st.text_input("ID del cementerio", value=registro.get("codigo_cementerio", ""), placeholder="Ej. CEM-001")
            pais = st.text_input("País", value="El Salvador", disabled=True)
            nombre_oficial = st.text_input("Nombre del cementerio", value=registro.get("nombre_oficial", ""))
            coordenadas = st.text_input("Coordenadas", value=registro.get("coordenadas", ""), placeholder="13.7010,-89.2240")

            idx_estado = ESTADOS.index(registro["estado"]) if registro.get("estado") and registro["estado"] in ESTADOS else 0
            estado = st.selectbox("Estado", ESTADOS, index=idx_estado)

            google_maps_url = st.text_input("Enlace Google Maps", value=registro.get("google_maps_url", ""), placeholder="https://...")

        with col2:
            enlace_principal = st.text_input("Enlace 1", value=registro.get("enlace_principal", ""), placeholder="https://...")
            enlace_2 = st.text_input("Enlace 2", value=registro.get("enlace_2", ""), placeholder="https://...")
            enlace_3 = st.text_input("Enlace 3", value=registro.get("enlace_3", ""), placeholder="https://...")
            foto_url = st.text_input("Enlace Fotografía", value=registro.get("foto_url", ""), placeholder="https://...")
            observaciones = st.text_area("Observaciones", value=registro.get("observaciones", ""), height=140)

        aprobado_qa = None
        comentarios_qa = None

        if st.session_state.user_role == "qa":
            st.divider()
            st.markdown("### 🔍 Control de Calidad")

            col_qa1, col_qa2 = st.columns(2)

            with col_qa1:
                aprobado_qa = st.checkbox(
                    "✅ Aprobar cementerio",
                    value=registro.get("aprobado_qa", False)
                )

            with col_qa2:
                comentarios_qa = st.text_area(
                    "Comentarios de QA",
                    value=registro.get("comentarios_qa", "") or "",
                    height=100
                )

        col_btn1, col_btn2 = st.columns(2)

        with col_btn1:
            actualizar = st.form_submit_button("Actualizar registro", use_container_width=True)

        with col_btn2:
            cancelar = st.form_submit_button("Cancelar", use_container_width=True)

        if cancelar:
            st.session_state.editando_id = None
            st.rerun()

        if actualizar:
            if not codigo_cementerio.strip():
                st.warning("Debe ingresar el ID del cementerio.")
                return

            if not nombre_oficial.strip():
                st.warning("Debe ingresar el nombre del cementerio.")
                return

            if not es_coordenada_valida(coordenadas):
                st.warning("Las coordenadas deben tener formato latitud,longitud.")
                return

            if not es_url_valida(google_maps_url):
                st.warning("El enlace de Google Maps debe empezar con http:// o https://")
                return

            if not es_url_valida(enlace_principal):
                st.warning("El Enlace 1 debe empezar con http:// o https://")
                return

            if not es_url_valida(enlace_2):
                st.warning("El Enlace 2 debe empezar con http:// o https://")
                return

            if not es_url_valida(enlace_3):
                st.warning("El Enlace 3 debe empezar con http:// o https://")
                return

            if not es_url_valida(foto_url):
                st.warning("El Enlace Fotografía debe empezar con http:// o https://")
                return

            datos = {
                "codigo_cementerio": codigo_cementerio.strip(),
                "nombre_oficial": nombre_oficial.strip(),
                "foto_url": foto_url.strip(),
                "coordenadas": coordenadas.strip(),
                "google_maps_url": google_maps_url.strip(),
                "estado": estado,
                "pais": "El Salvador",
                "departamento": departamento,
                "municipio": municipio,
                "distrito": distrito,
                "enlace_principal": enlace_principal.strip(),
                "enlace_2": enlace_2.strip(),
                "enlace_3": enlace_3.strip(),
                "observaciones": observaciones.strip(),
            }

            if st.session_state.user_role == "qa" and aprobado_qa is not None:
                datos["aprobado_qa"] = aprobado_qa
                datos["comentarios_qa"] = comentarios_qa.strip() if comentarios_qa else ""
                datos["revisado_por_qa"] = st.session_state.username
                datos["fecha_revision_qa"] = datetime.now().isoformat()

            ok, resultado = actualizar_cementerio(registro["id"], datos)

            if ok:
                st.success("Registro actualizado correctamente.")
                st.session_state.editando_id = None
                st.rerun()
            else:
                st.error(f"No se pudo actualizar: {resultado}")

def mostrar_registros():
    st.subheader("Ver datos capturados")

    datos = cargar_cementerios()

    if st.session_state.user_role == "colaborador":
        datos = [fila for fila in datos if str(fila.get("created_by")) == str(st.session_state.user)]

    if not datos:
        st.info("Todavía no hay registros para mostrar.")
        return

    for registro in datos:
        col1, col2, col3, col4 = st.columns([5, 1, 1, 1])

        with col1:
            qa_icon = "✅" if registro.get("aprobado_qa") else "⏳"
            st.write(f"{qa_icon} **{registro.get('nombre_oficial', 'Sin nombre')}** - {registro.get('departamento')} / {registro.get('municipio')} / {registro.get('distrito')}")

        with col2:
            st.write(f"_{registro.get('estado', 'N/A')}_")

        with col3:
            if st.button("✏️", key=f"edit_{registro['id']}", help="Editar"):
                st.session_state.editando_id = registro["id"]
                st.rerun()

        with col4:
            if st.session_state.user_role == "admin":
                if st.button("🗑️", key=f"del_{registro['id']}", help="Eliminar"):
                    if eliminar_cementerio(registro["id"]):
                        st.success("Registro eliminado")
                        st.rerun()
                    else:
                        st.error("Error al eliminar")

# ==============================
# PANEL QA - REVISIÓN
# ==============================
def mostrar_panel_revision_qa():
    st.subheader("🔍 Revisión de Cementerios")

    filtro = st.radio(
        "Filtrar por:",
        ["📋 Todos", "⏳ Pendientes de revisión", "✅ Aprobados"],
        horizontal=True
    )

    datos = cargar_cementerios()

    if filtro == "⏳ Pendientes de revisión":
        datos = [d for d in datos if not d.get("aprobado_qa", False)]
    elif filtro == "✅ Aprobados":
        datos = [d for d in datos if d.get("aprobado_qa", False)]

    if not datos:
        st.info("No hay cementerios en esta categoría.")
        return

    st.write(f"**{len(datos)} cementerios encontrados**")

    for registro in datos:
        with st.expander(f"{'✅' if registro.get('aprobado_qa') else '⏳'} {registro.get('nombre_oficial')} - {registro.get('departamento')}"):
            col1, col2 = st.columns(2)

            with col1:
                st.write(f"**ID:** {registro.get('codigo_cementerio')}")
                st.write(f"**Estado:** {registro.get('estado')}")
                st.write(f"**Departamento:** {registro.get('departamento')}")
                st.write(f"**Municipio:** {registro.get('municipio')}")
                st.write(f"**Distrito:** {registro.get('distrito')}")

            with col2:
                st.write(f"**Aprobado QA:** {'✅ Sí' if registro.get('aprobado_qa') else '⏳ Pendiente'}")
                if registro.get("revisado_por_qa"):
                    st.write(f"**Revisado por:** {registro.get('revisado_por_qa')}")
                if registro.get("comentarios_qa"):
                    st.write(f"**Comentarios QA:** {registro.get('comentarios_qa')}")

            if st.button("✏️ Revisar y editar", key=f"qa_edit_{registro['id']}"):
                st.session_state.editando_id = registro["id"]
                st.rerun()

# ==============================
# DASHBOARD
# ==============================
def mostrar_dashboard():
    st.subheader("📊 Dashboard")

    datos = cargar_cementerios()

    if st.session_state.user_role == "colaborador":
        datos = [fila for fila in datos if str(fila.get("created_by")) == str(st.session_state.user)]

    total = len(datos)
    certificados = len([x for x in datos if x.get("estado") == "Certificado"])
    aceptados = len([x for x in datos if x.get("estado") == "Aceptado"])
    provisionales = len([x for x in datos if x.get("estado") == "Provisional"])
    aprobados_qa = len([x for x in datos if x.get("aprobado_qa", False)])
    pendientes_qa = total - aprobados_qa

    st.markdown("### 📈 Resumen General")

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.markdown(f"""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        padding: 20px; border-radius: 14px; color: white; text-align: center;">
                <h2 style="margin:0; font-size: 2.4rem;">{total}</h2>
                <p style="margin:0; font-size: 1rem; opacity: 0.95;">Total Cementerios</p>
            </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown(f"""
            <div style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
                        padding: 20px; border-radius: 14px; color: white; text-align: center;">
                <h2 style="margin:0; font-size: 2.4rem;">{certificados}</h2>
                <p style="margin:0; font-size: 1rem; opacity: 0.95;">Certificados</p>
            </div>
        """, unsafe_allow_html=True)

    with col3:
        st.markdown(f"""
            <div style="background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
                        padding: 20px; border-radius: 14px; color: white; text-align: center;">
                <h2 style="margin:0; font-size: 2.4rem;">{aceptados}</h2>
                <p style="margin:0; font-size: 1rem; opacity: 0.95;">Aceptados</p>
            </div>
        """, unsafe_allow_html=True)

    with col4:
        st.markdown(f"""
            <div style="background: linear-gradient(135deg, #fa709a 0%, #fee140 100%);
                        padding: 20px; border-radius: 14px; color: white; text-align: center;">
                <h2 style="margin:0; font-size: 2.4rem;">{provisionales}</h2>
                <p style="margin:0; font-size: 1rem; opacity: 0.95;">Provisionales</p>
            </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    col5, col6 = st.columns(2)

    with col5:
        st.markdown(f"""
            <div style="background: linear-gradient(135deg, #43e97b 0%, #38f9d7 100%);
                        padding: 20px; border-radius: 14px; color: white; text-align: center;">
                <h2 style="margin:0; font-size: 2.4rem;">✅ {aprobados_qa}</h2>
                <p style="margin:0; font-size: 1rem; opacity: 0.95;">Aprobados QA</p>
            </div>
        """, unsafe_allow_html=True)

    with col6:
        st.markdown(f"""
            <div style="background: linear-gradient(135deg, #fccb90 0%, #d57eeb 100%);
                        padding: 20px; border-radius: 14px; color: white; text-align: center;">
                <h2 style="margin:0; font-size: 2.4rem;">⏳ {pendientes_qa}</h2>
                <p style="margin:0; font-size: 1rem; opacity: 0.95;">Pendientes QA</p>
            </div>
        """, unsafe_allow_html=True)

    st.divider()

    if total > 0:
        import pandas as pd
        import plotly.express as px

        st.markdown("### 📊 Visualizaciones")

        df = pd.DataFrame(datos)

        col_g1, col_g2 = st.columns(2)

        with col_g1:
            estados_count = df["estado"].value_counts().reset_index()
            estados_count.columns = ["Estado", "Cantidad"]

            fig_estados = px.pie(
                estados_count,
                values="Cantidad",
                names="Estado",
                title="Distribución por Estado",
                color_discrete_sequence=px.colors.qualitative.Set3
            )
            fig_estados.update_traces(textposition="inside", textinfo="percent+label")
            st.plotly_chart(fig_estados, use_container_width=True)

        with col_g2:
            qa_data = pd.DataFrame({
                "Estado QA": ["Aprobados", "Pendientes"],
                "Cantidad": [aprobados_qa, pendientes_qa]
            })

            fig_qa = px.bar(
                qa_data,
                x="Estado QA",
                y="Cantidad",
                title="Estado de Revisión QA",
                color="Estado QA",
                color_discrete_map={"Aprobados": "#43e97b", "Pendientes": "#fccb90"}
            )
            fig_qa.update_layout(showlegend=False)
            st.plotly_chart(fig_qa, use_container_width=True)

        dept_count = df["departamento"].value_counts().reset_index()
        dept_count.columns = ["Departamento", "Cantidad"]

        fig_dept = px.bar(
            dept_count,
            x="Departamento",
            y="Cantidad",
            title="Cementerios por Departamento",
            color="Cantidad",
            color_continuous_scale="Viridis"
        )
        fig_dept.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig_dept, use_container_width=True)

    st.divider()

    if total > 0:
        col_desc1, col_desc2, col_desc3 = st.columns([1, 2, 1])
        with col_desc2:
            archivo_excel = generar_excel(datos)
            st.download_button(
                label="📥 Descargar reporte en Excel",
                data=archivo_excel,
                file_name=f"cementerios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

# ==============================
# PANELES POR ROL
# ==============================
def mostrar_panel_admin():
    if st.session_state.editando_id:
        registro = cargar_cementerio_por_id(st.session_state.editando_id)
        if registro:
            mostrar_formulario_editar(registro)
        else:
            st.error("Registro no encontrado")
            st.session_state.editando_id = None
    else:
        tab1, tab2, tab3 = st.tabs(["➕ Nuevo registro", "📋 Ver datos capturados", "📊 Dashboard"])

        with tab1:
            mostrar_formulario_cementerio()

        with tab2:
            mostrar_registros()

        with tab3:
            mostrar_dashboard()

def mostrar_panel_colaborador():
    if st.session_state.editando_id:
        registro = cargar_cementerio_por_id(st.session_state.editando_id)
        if registro:
            mostrar_formulario_editar(registro)
        else:
            st.error("Registro no encontrado")
            st.session_state.editando_id = None
    else:
        tab1, tab2 = st.tabs(["➕ Nuevo registro", "📋 Mis registros"])

        with tab1:
            mostrar_formulario_cementerio()

        with tab2:
            mostrar_registros()

def mostrar_panel_qa():
    if st.session_state.editando_id:
        registro = cargar_cementerio_por_id(st.session_state.editando_id)
        if registro:
            mostrar_formulario_editar(registro)
        else:
            st.error("Registro no encontrado")
            st.session_state.editando_id = None
    else:
        tab1, tab2, tab3 = st.tabs(["🔍 Revisión QA", "📋 Todos los registros", "📊 Dashboard"])

        with tab1:
            mostrar_panel_revision_qa()

        with tab2:
            mostrar_registros()

        with tab3:
            mostrar_dashboard()

def mostrar_panel_consultor():
    tab1, tab2 = st.tabs(["📊 Dashboard", "📋 Registros"])

    with tab1:
        mostrar_dashboard()

    with tab2:
        mostrar_registros()

def mostrar_app():
    aplicar_estilos_app()

    rol_clase = f"badge-{st.session_state.user_role}"

    col1, col2 = st.columns([5, 1])
    with col1:
        st.markdown(
            f'<h1 style="margin-bottom:0.2rem;">Registro de cementerios <span class="rol-badge {rol_clase}">{st.session_state.user_role.upper()}</span></h1>',
            unsafe_allow_html=True
        )
        st.caption(f"👤 {st.session_state.user_nombre} (@{st.session_state.username})")

    with col2:
        if st.button("Cerrar sesión", use_container_width=True):
            st.session_state.user = None
            st.session_state.user_role = None
            st.session_state.user_nombre = None
            st.session_state.username = None
            st.session_state.editando_id = None
            st.rerun()

    st.divider()

    if st.session_state.user_role == "admin":
        mostrar_panel_admin()
    elif st.session_state.user_role == "colaborador":
        mostrar_panel_colaborador()
    elif st.session_state.user_role == "qa":
        mostrar_panel_qa()
    elif st.session_state.user_role == "consultor":
        mostrar_panel_consultor()
    else:
        st.warning("Rol no reconocido.")

# ==============================
# FLUJO PRINCIPAL
# ==============================
if st.session_state.user is None:
    mostrar_login()
    st.stop()
else:
    mostrar_app()